"""
Email Campaign Analyzer
-------------------------
Analyzes email campaign CSV exports from any platform (Mailchimp,
Klaviyo, ConvertKit, etc.) by auto-detecting column names.

Calculates open rate, CTR, unsubscribe rate, bounce rate, and trends.
Exports a multi-sheet Excel report with embedded charts.

Usage:
    python email_analyzer.py --file email_campaigns.csv
    python email_analyzer.py --file email_campaigns.csv --output report.xlsx
    python email_analyzer.py --file email_campaigns.csv --map      (interactive column mapping)
"""

import argparse
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─── COLUMN DETECTION ─────────────────────────────────────────────────────────

# Maps canonical field names to known column name variants across platforms
COLUMN_ALIASES = {
    "campaign_name": ["campaign_name", "campaign", "subject", "email_name", "name", "title"],
    "send_date":     ["send_date", "date_sent", "sent_date", "date", "send_time", "scheduled_date"],
    "emails_sent":   ["emails_sent", "sent", "total_sent", "recipients", "audience_size", "num_recipients"],
    "emails_opened": ["emails_opened", "opened", "unique_opens", "opens", "total_opens"],
    "unique_clicks": ["unique_clicks", "clicks", "click_throughs", "total_clicks", "unique_link_clicks"],
    "unsubscribes":  ["unsubscribes", "unsubscribed", "unsubs", "opt_outs", "total_unsubscribes"],
    "bounces":       ["bounces", "total_bounces", "bounce_count", "hard_bounces", "soft_bounces"],
}


def detect_columns(df_cols: list) -> dict:
    """
    Auto-detect which CSV columns map to which canonical fields.
    Returns {canonical_name: actual_col_name}.
    """
    cols_lower = {c.lower().strip(): c for c in df_cols}
    mapping = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in cols_lower:
                mapping[canonical] = cols_lower[alias]
                break
    return mapping


def interactive_map(df_cols: list, mapping: dict) -> dict:
    """Let user manually map any undetected required columns."""
    required = ["emails_sent", "emails_opened", "unique_clicks"]
    missing = [r for r in required if r not in mapping]

    if not missing:
        return mapping

    print("\n⚠️  Could not auto-detect the following columns:")
    print(f"   Available columns: {', '.join(df_cols)}\n")

    for field in missing:
        while True:
            val = input(f"  Which column is '{field}'? (or press Enter to skip): ").strip()
            if val == "":
                break
            if val in df_cols:
                mapping[field] = val
                break
            print(f"  '{val}' not found. Try again.")
    return mapping


# ─── LOAD & CALCULATE ─────────────────────────────────────────────────────────

def load_and_calculate(filepath: str, mapping: dict) -> pd.DataFrame:
    """Load CSV, rename columns to canonical names, calculate rates."""
    df = pd.read_csv(filepath)

    # Rename detected columns to canonical names
    reverse_map = {v: k for k, v in mapping.items()}
    df = df.rename(columns=reverse_map)

    # Parse date if present
    if "send_date" in df.columns:
        df["send_date"] = pd.to_datetime(df["send_date"], errors="coerce")
        df = df.sort_values("send_date").reset_index(drop=True)
        df["month"] = df["send_date"].dt.to_period("M").astype(str)

    # Calculate rates (guard against division by zero)
    sent = df["emails_sent"].replace(0, pd.NA)
    opened = df.get("emails_opened", pd.Series([pd.NA] * len(df))).replace(0, pd.NA)

    df["open_rate"]   = (df["emails_opened"] / sent * 100).round(2)
    df["ctr"]         = (df["unique_clicks"] / sent * 100).round(2)
    df["click_to_open"] = (df["unique_clicks"] / opened * 100).round(2)

    if "unsubscribes" in df.columns:
        df["unsub_rate"] = (df["unsubscribes"] / sent * 100).round(3)
    if "bounces" in df.columns:
        df["bounce_rate"] = (df["bounces"] / sent * 100).round(3)

    return df


# ─── ANALYSIS ─────────────────────────────────────────────────────────────────

def monthly_trends(df: pd.DataFrame) -> pd.DataFrame:
    if "month" not in df.columns:
        return pd.DataFrame()
    agg = df.groupby("month").agg(
        campaigns=("emails_sent", "count"),
        avg_open_rate=("open_rate", "mean"),
        avg_ctr=("ctr", "mean"),
        total_sent=("emails_sent", "sum"),
        total_opens=("emails_opened", "sum"),
        total_clicks=("unique_clicks", "sum"),
    ).round(2).reset_index()
    return agg


def top_campaigns(df: pd.DataFrame, by: str = "open_rate", n: int = 5) -> pd.DataFrame:
    cols = ["campaign_name", "send_date", "emails_sent", "open_rate", "ctr"]
    if "unsub_rate" in df.columns:
        cols.append("unsub_rate")
    available = [c for c in cols if c in df.columns]
    return df.nlargest(n, by)[available].reset_index(drop=True)


def bottom_campaigns(df: pd.DataFrame, by: str = "open_rate", n: int = 5) -> pd.DataFrame:
    cols = ["campaign_name", "send_date", "emails_sent", "open_rate", "ctr"]
    available = [c for c in cols if c in df.columns]
    return df.nsmallest(n, by)[available].reset_index(drop=True)


def overall_stats(df: pd.DataFrame) -> dict:
    return {
        "Total Campaigns": len(df),
        "Total Emails Sent": int(df["emails_sent"].sum()),
        "Avg Open Rate (%)": round(df["open_rate"].mean(), 2),
        "Avg CTR (%)": round(df["ctr"].mean(), 2),
        "Best Open Rate (%)": round(df["open_rate"].max(), 2),
        "Worst Open Rate (%)": round(df["open_rate"].min(), 2),
        "Avg Unsub Rate (%)": round(df["unsub_rate"].mean(), 3) if "unsub_rate" in df.columns else "N/A",
        "Avg Bounce Rate (%)": round(df["bounce_rate"].mean(), 3) if "bounce_rate" in df.columns else "N/A",
        "Date Range": f"{df['send_date'].min().date()} → {df['send_date'].max().date()}" if "send_date" in df.columns else "N/A",
        "Analyzed On": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }


# ─── TERMINAL SUMMARY ─────────────────────────────────────────────────────────

def print_summary(df: pd.DataFrame):
    stats = overall_stats(df)
    print("\n" + "=" * 55)
    print("  EMAIL CAMPAIGN ANALYSIS")
    print("=" * 55)
    for k, v in stats.items():
        print(f"  {k:<28} {v}")

    print("\n── Top 5 Campaigns by Open Rate ──")
    print(top_campaigns(df).to_string(index=False))

    print("\n── Bottom 5 Campaigns by Open Rate ──")
    print(bottom_campaigns(df).to_string(index=False))
    print()


# ─── EXCEL HELPERS ────────────────────────────────────────────────────────────

COLOR_HEADER  = "365349"
COLOR_SUB     = "435066"
COLOR_GREEN   = "C6EFCE"
COLOR_YELLOW  = "FFEB9C"
COLOR_RED     = "FFC7CE"


def style_header(ws, num_cols: int, color: str = COLOR_HEADER):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")


def auto_fit(ws, max_width: int = 50):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, max_width)


def color_rate_column(ws, col_idx: int, low: float, high: float, start_row: int = 2):
    """Green if >= high, yellow if between low and high, red if < low."""
    for row in ws.iter_rows(min_row=start_row, min_col=col_idx, max_col=col_idx):
        for cell in row:
            if cell.value is None:
                continue
            try:
                val = float(cell.value)
                if val >= high:
                    color = COLOR_GREEN
                elif val >= low:
                    color = COLOR_YELLOW
                else:
                    color = COLOR_RED
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            except (TypeError, ValueError):
                pass


# ─── CHART BUILDERS ───────────────────────────────────────────────────────────

def add_open_rate_trend_chart(wb, monthly_df: pd.DataFrame):
    """Line chart: monthly avg open rate trend on the Monthly Trends sheet."""
    ws = wb["Monthly Trends"]

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    open_col = headers.index("avg_open_rate") + 1
    month_col = headers.index("month") + 1
    num_rows = len(monthly_df) + 1

    chart = LineChart()
    chart.title = "Avg Open Rate Trend by Month (%)"
    chart.y_axis.title = "Open Rate (%)"
    chart.x_axis.title = "Month"
    chart.style = 10
    chart.width = 22
    chart.height = 14

    data = Reference(ws, min_col=open_col, min_row=1, max_row=num_rows)
    cats = Reference(ws, min_col=month_col, min_row=2, max_row=num_rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "I2")


def add_open_vs_ctr_chart(wb, monthly_df: pd.DataFrame):
    """Bar chart: open rate vs CTR side by side per month."""
    ws = wb["Monthly Trends"]

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    open_col = headers.index("avg_open_rate") + 1
    ctr_col  = headers.index("avg_ctr") + 1
    month_col = headers.index("month") + 1
    num_rows = len(monthly_df) + 1

    chart = BarChart()
    chart.type = "col"
    chart.title = "Open Rate vs CTR by Month (%)"
    chart.y_axis.title = "%"
    chart.x_axis.title = "Month"
    chart.style = 10
    chart.width = 22
    chart.height = 14

    data_open = Reference(ws, min_col=open_col, min_row=1, max_row=num_rows)
    data_ctr  = Reference(ws, min_col=ctr_col,  min_row=1, max_row=num_rows)
    cats = Reference(ws, min_col=month_col, min_row=2, max_row=num_rows)

    chart.add_data(data_open, titles_from_data=True)
    chart.add_data(data_ctr,  titles_from_data=True)
    chart.set_categories(cats)

    # Place below the trend chart
    row_offset = len(monthly_df) + 20
    ws.add_chart(chart, f"I{row_offset}")


def add_volume_chart(wb, monthly_df: pd.DataFrame):
    """Bar chart: total emails sent per month."""
    ws = wb["Monthly Trends"]

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    sent_col  = headers.index("total_sent") + 1
    month_col = headers.index("month") + 1
    num_rows  = len(monthly_df) + 1

    chart = BarChart()
    chart.type = "col"
    chart.title = "Total Emails Sent per Month"
    chart.y_axis.title = "Emails Sent"
    chart.style = 10
    chart.width = 22
    chart.height = 12

    data = Reference(ws, min_col=sent_col, min_row=1, max_row=num_rows)
    cats = Reference(ws, min_col=month_col, min_row=2, max_row=num_rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    row_offset = len(monthly_df) + 40
    ws.add_chart(chart, f"I{row_offset}")


# ─── EXCEL EXPORT ─────────────────────────────────────────────────────────────

def export_excel(df: pd.DataFrame, monthly_df: pd.DataFrame, output_file: str):
    stats = overall_stats(df)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

        # Sheet 1: All campaigns
        display_cols = [c for c in [
            "campaign_name", "send_date", "emails_sent", "emails_opened",
            "open_rate", "unique_clicks", "ctr", "click_to_open",
            "unsubscribes", "unsub_rate", "bounces", "bounce_rate"
        ] if c in df.columns]
        df[display_cols].to_excel(writer, sheet_name="All Campaigns", index=False)

        # Sheet 2: Monthly trends
        if not monthly_df.empty:
            monthly_df.to_excel(writer, sheet_name="Monthly Trends", index=False)

        # Sheet 3: Top campaigns
        top_campaigns(df, n=10).to_excel(writer, sheet_name="Top 10 Campaigns", index=False)

        # Sheet 4: Bottom campaigns
        bottom_campaigns(df, n=10).to_excel(writer, sheet_name="Bottom 10 Campaigns", index=False)

        # Sheet 5: Summary
        pd.DataFrame(list(stats.items()), columns=["Metric", "Value"]).to_excel(
            writer, sheet_name="Summary", index=False
        )

    # Post-process: styling + charts
    wb = load_workbook(output_file)

    # Style all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        style_header(ws, ws.max_column)
        auto_fit(ws)
        ws.freeze_panes = "A2"

    # Color-code open_rate and ctr columns on All Campaigns sheet
    ws_all = wb["All Campaigns"]
    headers = [ws_all.cell(row=1, column=c).value for c in range(1, ws_all.max_column + 1)]
    if "open_rate" in headers:
        color_rate_column(ws_all, headers.index("open_rate") + 1, low=20.0, high=30.0)
    if "ctr" in headers:
        color_rate_column(ws_all, headers.index("ctr") + 1, low=1.5, high=3.0)

    # Add charts
    if not monthly_df.empty:
        add_open_rate_trend_chart(wb, monthly_df)
        add_open_vs_ctr_chart(wb, monthly_df)
        add_volume_chart(wb, monthly_df)

    wb.save(output_file)
    print(f"  ✅ Report saved: {output_file}")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Analyze email campaign CSV exports from any platform.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python email_analyzer.py --file campaigns.csv
  python email_analyzer.py --file campaigns.csv --output report.xlsx
  python email_analyzer.py --file campaigns.csv --map
        """
    )
    parser.add_argument("--file", type=str, required=True, help="Path to CSV file")
    parser.add_argument("--output", type=str, default="email_report.xlsx", help="Output Excel filename")
    parser.add_argument("--map", action="store_true", help="Interactively map undetected columns")

    args = parser.parse_args()

    if not Path(args.file).exists():
        print(f"❌ File not found: {args.file}")
        sys.exit(1)

    print("\n📧 Email Campaign Analyzer")
    print("-" * 40)

    # Load CSV and detect columns
    df_raw = pd.read_csv(args.file)
    print(f"  📂 Loaded: {args.file} ({len(df_raw)} rows)")
    print(f"  🔍 Columns found: {', '.join(df_raw.columns.tolist())}")

    mapping = detect_columns(df_raw.columns.tolist())
    detected = list(mapping.keys())
    print(f"  ✅ Auto-detected: {', '.join(detected)}")

    missing_required = [r for r in ["emails_sent", "emails_opened", "unique_clicks"] if r not in mapping]
    if missing_required:
        print(f"  ⚠️  Could not detect: {', '.join(missing_required)}")
        if args.map:
            mapping = interactive_map(df_raw.columns.tolist(), mapping)
        else:
            print("  💡 Re-run with --map to manually assign columns.")

    # Analyze
    df = load_and_calculate(args.file, mapping)
    monthly_df = monthly_trends(df)

    # Output
    print_summary(df)
    output = args.output if args.output.endswith(".xlsx") else args.output + ".xlsx"
    export_excel(df, monthly_df, output)


if __name__ == "__main__":
    main()
